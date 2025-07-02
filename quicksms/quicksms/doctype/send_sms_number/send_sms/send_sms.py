import frappe
import os
import base64
import requests
import chardet
import openpyxl
import csv
import datetime
from datetime import time
from frappe.model.document import Document
from frappe.utils import get_datetime, now_datetime
from frappe.utils.background_jobs import enqueue
from quicksms.sms_service import SmsService
from frappe.utils import strip_html
import html
from frappe import _
from requests.auth import HTTPBasicAuth
import re


class SendSMS(Document):
    def before_save(self):
        summary = []
        for row in self.numbers[:5]:  
            phone = row.phone_number or "-"
            summary.append(phone)
        self.numbers_summary = ", ".join(summary)


@frappe.whitelist()
def process_and_send_sms(docname):
    doc = frappe.get_doc("Send SMS", docname)

    if not doc.numbers:
        frappe.throw("No numbers found in the table.")

    # If schedule_sms is checked, we will schedule the SMS
    if doc.schedule_sms and doc.scheduled_date and doc.scheduled_time:
        # Scheduling SMS logic
        if isinstance(doc.scheduled_time, datetime.timedelta):
            hours, remainder = divmod(doc.scheduled_time.seconds, 3600)
            minutes, seconds = divmod(remainder, 60)
            scheduled_time = time(hour=hours, minute=minutes, second=seconds)
        else:
            scheduled_time = doc.scheduled_time

        scheduled_datetime = datetime.datetime.combine(doc.scheduled_date, scheduled_time)

        frappe.logger().info(f"[SMS] Scheduled for: {scheduled_datetime}")
        frappe.logger().info(f"[SMS] Current time: {now_datetime()}")

        if scheduled_datetime > now_datetime():
            # Update SMS status to Scheduled and save
            doc.status = "Scheduled"
            doc.save()

            enqueue(
                method="quicksms.quicksms.doctype.send_sms.send_sms.send_sms_job",
                queue="long",
                at=get_datetime(scheduled_datetime),
                docname=docname
            )

            frappe.logger().info(f"[SMS] Enqueued SMS job for {docname} at {scheduled_datetime}")
            return f"SMS scheduled for {scheduled_datetime.strftime('%Y-%m-%d %H:%M:%S')}."

    # If schedule_sms is not checked or the scheduled time is in the past, send SMS now
    return send_sms_now(doc)




def send_sms_job(docname):
    frappe.logger().info(f"[SMS JOB STARTED] For docname: {docname}")
    try:
        doc = frappe.get_doc("Send SMS", docname)
        return send_sms_now(doc)
    except Exception:
        frappe.log_error(frappe.get_traceback(), f"[SMS JOB ERROR] Failed for doc: {docname}")
        try:
            # Try to mark status as failed if doc is found
            doc = frappe.get_doc("Send SMS", docname)
            doc.status = "Failed"
            doc.save()
        except Exception:
            frappe.logger().error(f"Unable to mark failed status for doc: {docname}")



def send_sms_now(doc):
    frappe.logger().info(f"[SMS SENDING] docname: {doc.name}, Total Numbers: {len(doc.numbers)}")
    template = frappe.get_doc("SMS Templates", doc.sms_for)
    raw_content = template.content or "Hello {code}, your appointment is on {date}"
    clean_html = strip_html(raw_content)
    content = html.unescape(clean_html)

    settings = frappe.get_doc("Quick SMS Settings")
    app_id = settings.app_id.strip()
    app_secret = settings.get_password("app_secret").strip()
    base_url = settings.base_url.strip().rstrip("/")
    sender = getattr(settings, "sender_name", "").strip() or "MAMPCO"

    auth_string = f"{app_id}:{app_secret}"
    auth_header = "Basic " + base64.b64encode(auth_string.encode()).decode("utf-8")
    headers = {
        "Authorization": auth_header,
        "Accept": "application/json",
        "Content-Type": "application/json"
    }

    success_count = 0

    for row in doc.numbers:
        phone = row.phone_number
        code = row.code
        message = content.replace("{code}", code).replace("{date}", str(doc.date))

        payload = {
            "messages": [
                {
                    "text": message,
                    "numbers": [phone],
                    "sender": sender
                }
            ]
        }

        try:
            resp = requests.post(f"{base_url}/account/area/sms/send", headers=headers, json=payload, timeout=10)
            frappe.log_error(title=f"SMS Sent to {phone}", message=f"Status Code: {resp.status_code}\nResponse: {resp.text}")

            if resp.status_code == 200 and resp.json().get("code") == 200:
                success_count += 1
                row.status = "Sent"  # ✅ Mark individual row as Sent
            else:
                row.status = "Failed"  # ❌ Mark as Failed
                frappe.log_error(str(resp.json()), f"SMS Failed for {phone}")
        except Exception:
            row.status = "Failed"  # ❌ Catch exception also as Failed
            frappe.log_error(frappe.get_traceback(), f"Exception for {phone}")

    doc.status = "Sent" if success_count > 0 else "Failed"
    doc.save(ignore_permissions=True)

    frappe.publish_realtime("sms_sent", {"docname": doc.name})
    return f"{success_count} SMS sent successfully. {len(doc.numbers) - success_count} SMS failed."


@frappe.whitelist()
def upload_numbers_to_child_table(docname, file_url):
    doc = frappe.get_doc("Send SMS", docname)
    file_doc = frappe.get_doc("File", {"file_url": file_url})
    file_path = file_doc.get_full_path()
    ext = os.path.splitext(file_path)[1].lower()

    new_rows = []

    try:
        if ext in ['.xlsx', '.xls']:
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                code = str(row[0]).strip() if row[0] else ""
                phone = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                if code and phone:
                    new_rows.append({"code": code, "phone_number": phone})

        elif ext == '.csv':
            with open(file_path, 'rb') as f:
                raw = f.read()
                encoding = chardet.detect(raw)['encoding'] or 'utf-8'

            with open(file_path, 'r', encoding=encoding, errors='ignore') as f:
                reader = csv.reader(f)
                next(reader, None)
                for row in reader:
                    if len(row) >= 2 and row[0].strip() and row[1].strip():
                        new_rows.append({"code": row[0].strip(), "phone_number": row[1].strip()})

    except Exception:
        frappe.log_error(frappe.get_traceback(), "Upload Failed")
        frappe.throw("Failed to process the uploaded file.")

    for r in new_rows:
        doc.append("numbers", r)

    doc.save()
    return f"{len(new_rows)} numbers added successfully."


def extract_phone_numbers(file_path):
    ext = os.path.splitext(file_path)[1].lower()
    numbers = []

    try:
        if ext in ['.xlsx', '.xls']:
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                phone = str(row[1]).strip() if row and len(row) > 1 else None
                if phone:
                    numbers.append(phone)

        elif ext == '.csv':
            with open(file_path, 'rb') as f:
                raw = f.read()
                encoding = chardet.detect(raw)['encoding'] or 'utf-8'

            with open(file_path, 'r', encoding=encoding, errors='ignore') as f:
                reader = csv.reader(f)
                next(reader, None)
                for row in reader:
                    if len(row) >= 2 and row[1].strip():
                        numbers.append(row[1].strip())

    except Exception:
        frappe.log_error(frappe.get_traceback(), "Failed to extract phone numbers")

    return numbers


@frappe.whitelist()
def fetch_sms_balance_for_listview():
    try:
        sms_service = SmsService()
        result = sms_service.get_balance()

        if result.get("error"):
            return {"error": result["error"]}

        return {
            "success": True,
            "balance": result.get("total_balance", 0)
        }

    except Exception as e:
        return {"error": str(e)}



def process_due_scheduled_sms():
    now = now_datetime()

    # Get all scheduled SMS entries where:
    # - schedule_sms is 1 (checked)
    # - sms_status is "Scheduled"
    # - scheduled datetime <= now
    docs = frappe.get_all("Send SMS", 
        filters={
            "schedule_sms": 1,
            "status": "Scheduled"
        }, 
        fields=["name", "scheduled_date", "scheduled_time"]
    )

    for d in docs:
        try:
            # Build full scheduled datetime
            if isinstance(d.scheduled_time, datetime.timedelta):
                hours, remainder = divmod(d.scheduled_time.seconds, 3600)
                minutes, seconds = divmod(remainder, 60)
                scheduled_time = time(hour=hours, minute=minutes, second=seconds)
            else:
                scheduled_time = d.scheduled_time

            scheduled_datetime = datetime.datetime.combine(d.scheduled_date, scheduled_time)

            # If scheduled time is due or passed, send it
            if scheduled_datetime <= now:
                frappe.logger().info(f"[SCHEDULED SMS EXECUTION] Sending SMS for doc: {d.name}")
                doc = frappe.get_doc("Send SMS", d.name)
                send_sms_now(doc)

        except Exception:
            frappe.log_error(frappe.get_traceback(), f"[Scheduled SMS Failed] Doc: {d.name}")




@frappe.whitelist()
def fetch_sms_history(from_date=None, to_date=None, number=None, approve_status=None):
    def format_date(date_str):
        try:
            return datetime.datetime.strptime(date_str, "%Y-%m-%d").strftime("%d/%m/%Y")
        except Exception:
            return date_str

    settings = frappe.get_doc("Quick SMS Settings")
    app_id = settings.app_id.strip()
    app_secret = settings.get_password("app_secret").strip()

    auth_string = f"{app_id}:{app_secret}"
    auth_header = "Basic " + base64.b64encode(auth_string.encode()).decode("utf-8")

    headers = {
        "Authorization": auth_header,
        "Accept": "application/json",
        "Content-Type": "application/json"
    }

    base_url = "https://api-sms.4jawaly.com/api/v1/account/area/sms/archive"
    params = {
        "take": 100,
        "page": 1
    }

    if from_date:
        params["created_at_from"] = format_date(from_date)
    if to_date:
        params["created_at_to"] = format_date(to_date)
    if number:
        params["number"] = number
    if approve_status and approve_status != "All":
        params["approve_status"] = approve_status.strip().split(" ")[0]

    try:
        all_data = []
        while True:
            response = requests.get(base_url, headers=headers, params=params, timeout=20)
            response.raise_for_status()

            json_data = response.json()
            page_data = json_data.get("paginate", {}).get("data", [])
            all_data.extend(page_data)

            next_url = json_data.get("paginate", {}).get("next_page_url")
            if not next_url:
                break

            # Move to next page
            params["page"] += 1

        return {
            "status": "success",
            "data": all_data
        }

    except requests.exceptions.RequestException as e:
        err_msg = getattr(e.response, 'text', str(e))
        frappe.log_error(frappe.get_traceback(), "Fetch SMS History Error")
        return {
            "status": "error",
            "message": f"Failed to fetch SMS history: {err_msg}"
        }